"""Парсер для Excel и CSV файлов с товарами."""

import csv
import io
from decimal import Decimal, InvalidOperation
from typing import Any, NamedTuple

from src.core.logging import get_logger

logger = get_logger(__name__)


class ProductRow(NamedTuple):
    """Строка с данными товара из файла."""

    name: str
    description: str | None
    price: Decimal
    sizes: list[str]
    category_name: str
    photo_url: str | None
    row_number: int


class ParseError(NamedTuple):
    """Ошибка парсинга строки."""

    row_number: int
    field: str
    error: str


class ParseResult(NamedTuple):
    """Результат парсинга файла."""

    products: list[ProductRow]
    errors: list[ParseError]


class ExcelParser:
    """Парсер для файлов с товарами."""

    # Ожидаемые заголовки (case-insensitive)
    EXPECTED_HEADERS = {
        "название",
        "описание",
        "цена",
        "размеры",
        "категория",
        "фото",
    }

    # Альтернативные названия столбцов
    HEADER_ALIASES = {
        "название": ["название", "name", "product", "товар"],
        "описание": ["описание", "description", "desc"],
        "цена": ["цена", "price", "стоимость"],
        "размеры": ["размеры", "sizes", "size", "размер"],
        "категория": ["категория", "category", "cat"],
        "фото": ["фото", "photo", "image", "url", "photo_url", "картинка"],
    }

    @classmethod
    def parse_csv(cls, file_content: bytes) -> ParseResult:
        """Парсить CSV файл.

        Args:
            file_content: Содержимое CSV файла

        Returns:
            Результат парсинга
        """
        try:
            # Пробуем разные кодировки
            for encoding in ["utf-8", "cp1251", "latin-1"]:
                try:
                    text = file_content.decode(encoding)
                    break
                except UnicodeDecodeError:
                    continue
            else:
                return ParseResult(
                    products=[],
                    errors=[
                        ParseError(0, "file", "Не удалось определить кодировку файла")
                    ],
                )

            # Определяем разделитель
            sniffer = csv.Sniffer()
            try:
                delimiter = sniffer.sniff(text[:1024]).delimiter
            except csv.Error:
                delimiter = ";"  # По умолчанию

            # Парсим CSV
            reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)

            return cls._parse_rows(reader)

        except Exception as e:
            logger.error(f"CSV parsing failed: {e}", exc_info=True)
            return ParseResult(
                products=[],
                errors=[ParseError(0, "file", f"Ошибка чтения файла: {str(e)}")],
            )

    @classmethod
    def parse_excel(cls, file_content: bytes) -> ParseResult:
        """Парсить Excel файл.

        Args:
            file_content: Содержимое Excel файла

        Returns:
            Результат парсинга
        """
        try:
            import openpyxl
        except ImportError:
            return ParseResult(
                products=[],
                errors=[
                    ParseError(
                        0,
                        "file",
                        "Библиотека openpyxl не установлена. Установите: pip install openpyxl",
                    )
                ],
            )

        try:
            # Загружаем Excel файл
            wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True)
            ws = wb.active

            # Читаем заголовки из первой строки
            headers = []
            for cell in ws[1]:
                value = cell.value
                headers.append(str(value).strip().lower() if value else "")

            # Создаем список словарей для каждой строки
            rows = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                row_dict = {}
                for idx, cell in enumerate(row):
                    if idx < len(headers):
                        row_dict[headers[idx]] = cell.value
                row_dict["_row_number"] = row_idx
                rows.append(row_dict)

            return cls._parse_rows(rows)

        except Exception as e:
            logger.error(f"Excel parsing failed: {e}", exc_info=True)
            return ParseResult(
                products=[],
                errors=[ParseError(0, "file", f"Ошибка чтения Excel: {str(e)}")],
            )

    @classmethod
    def _parse_rows(cls, rows: Any) -> ParseResult:
        """Парсить строки из файла.

        Args:
            rows: Итератор по строкам (dict)

        Returns:
            Результат парсинга
        """
        products = []
        errors = []

        # Мапируем заголовки
        row_list = list(rows)
        if not row_list:
            return ParseResult(
                products=[],
                errors=[ParseError(0, "file", "Файл пуст или нет данных")],
            )

        # Получаем заголовки из первой строки
        first_row = row_list[0]
        header_mapping = cls._map_headers(first_row.keys())

        if not header_mapping:
            return ParseResult(
                products=[],
                errors=[
                    ParseError(
                        0,
                        "file",
                        f"Не найдены обязательные столбцы. Ожидаются: {', '.join(cls.EXPECTED_HEADERS)}",
                    )
                ],
            )

        # Парсим каждую строку
        for row_idx, row in enumerate(row_list, start=2):
            try:
                # Извлекаем номер строки если есть
                row_number = row.get("_row_number", row_idx)

                # Получаем значения
                name = cls._get_value(row, header_mapping.get("название"), "")
                description = cls._get_value(row, header_mapping.get("описание"))
                price_str = cls._get_value(row, header_mapping.get("цена"), "")
                sizes_str = cls._get_value(row, header_mapping.get("размеры"), "")
                category = cls._get_value(row, header_mapping.get("категория"), "")
                photo_url = cls._get_value(row, header_mapping.get("фото"))

                # Валидация и преобразование
                row_errors = []

                # Название (обязательно)
                if not name or not str(name).strip():
                    row_errors.append(
                        ParseError(row_number, "название", "Не указано название")
                    )
                    continue

                name = str(name).strip()

                # Цена (обязательно)
                try:
                    if not price_str:
                        raise ValueError("Цена не указана")

                    # Убираем пробелы и заменяем запятую на точку
                    price_clean = str(price_str).replace(" ", "").replace(",", ".")
                    price = Decimal(price_clean)

                    if price <= 0:
                        raise ValueError("Цена должна быть положительной")

                except (InvalidOperation, ValueError) as e:
                    row_errors.append(
                        ParseError(
                            row_number, "цена", f"Некорректная цена: {str(e)}"
                        )
                    )
                    continue

                # Размеры (обязательно)
                if not sizes_str:
                    row_errors.append(
                        ParseError(row_number, "размеры", "Не указаны размеры")
                    )
                    continue

                # Парсим размеры (через запятую, точку с запятой или пробел)
                sizes_raw = str(sizes_str).replace(";", ",").replace("|", ",")
                sizes = [s.strip().upper() for s in sizes_raw.split(",") if s.strip()]

                if not sizes:
                    row_errors.append(
                        ParseError(row_number, "размеры", "Размеры не распознаны")
                    )
                    continue

                # Категория (обязательно)
                if not category or not str(category).strip():
                    row_errors.append(
                        ParseError(row_number, "категория", "Не указана категория")
                    )
                    continue

                category = str(category).strip()

                # Описание (опционально)
                description = str(description).strip() if description else None

                # URL фото (опционально)
                photo_url = str(photo_url).strip() if photo_url else None

                # Если есть ошибки - добавляем и пропускаем
                if row_errors:
                    errors.extend(row_errors)
                    continue

                # Создаём ProductRow
                product_row = ProductRow(
                    name=name,
                    description=description,
                    price=price,
                    sizes=sizes,
                    category_name=category,
                    photo_url=photo_url,
                    row_number=row_number,
                )

                products.append(product_row)

            except Exception as e:
                logger.error(f"Error parsing row {row_idx}: {e}", exc_info=True)
                errors.append(
                    ParseError(row_idx, "строка", f"Ошибка обработки: {str(e)}")
                )

        return ParseResult(products=products, errors=errors)

    @classmethod
    def _map_headers(cls, headers: Any) -> dict[str, str]:
        """Мапировать заголовки на стандартные названия.

        Args:
            headers: Заголовки из файла

        Returns:
            Словарь {стандартное_имя: имя_в_файле}
        """
        mapping = {}
        headers_lower = {h.lower().strip(): h for h in headers if h}

        for standard_name, aliases in cls.HEADER_ALIASES.items():
            for alias in aliases:
                if alias in headers_lower:
                    mapping[standard_name] = headers_lower[alias]
                    break

        return mapping

    @classmethod
    def _get_value(cls, row: dict, key: str | None, default: Any = None) -> Any:
        """Получить значение из строки по ключу.

        Args:
            row: Строка данных
            key: Ключ
            default: Значение по умолчанию

        Returns:
            Значение или default
        """
        if not key:
            return default

        value = row.get(key)
        if value is None or (isinstance(value, str) and not value.strip()):
            return default

        return value
